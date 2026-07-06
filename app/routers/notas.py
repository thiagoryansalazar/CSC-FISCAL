import csv
import io
import json
import os
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response, FileResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.services.nota_service import NotaService
from app.schemas.nota_schema import NotaFiscalCreate, NotaFiscalUpdate, NotaFiscalResponse, ItemNotaResponse, HistoricoResponse
from app.models.item_nota import ItemNota


router = APIRouter(prefix='/api/notas', tags=['notas'])

@router.get('')
async def listar_notas(
    status: str = None,
    search: str = None,
    cnpj: str = None,
    data_inicio: str = None,
    data_fim: str = None,
    offset: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    svc = NotaService(db)
    notas = await svc.listar(status, search, cnpj, data_inicio, data_fim, offset, limit)
    return [NotaFiscalResponse.model_validate(n) for n in notas]

@router.get('/export')
async def exportar_notas(
    formato: str = Query('json', pattern='^(json|csv|xlsx)$'),
    status: str = None,
    search: str = None,
    cnpj: str = None,
    data_inicio: str = None,
    data_fim: str = None,
    db: AsyncSession = Depends(get_db),
):
    svc = NotaService(db)
    notas = await svc.listar(status, search, cnpj, data_inicio, data_fim, 0, 10000)
    dados = [NotaFiscalResponse.model_validate(n).model_dump() for n in notas]

    if formato == 'json':
        return Response(
            content=json.dumps(dados, ensure_ascii=False, indent=2, default=str),
            media_type='application/json',
            headers={'Content-Disposition': 'attachment; filename=notas_fiscais.json'},
        )

    if formato == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['id', 'numero', 'serie', 'fornecedor', 'cnpj', 'valor', 'status', 'data_emissao', 'chave_acesso'])
        for n in dados:
            writer.writerow([
                n['id'], n['numero'], n['serie'], n['nome_fornecedor'],
                n['cnpj_emitente'], n['valor_total'], n['status'], n['data_emissao'], n['chave_acesso'],
            ])
        return Response(
            content='\ufeff' + output.getvalue(),
            media_type='text/csv; charset=utf-8',
            headers={'Content-Disposition': 'attachment; filename=notas_fiscais.csv'},
        )

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notas'
    cab = ['ID', 'Número', 'Série', 'Fornecedor', 'CNPJ', 'Valor', 'Status', 'Data Emissão', 'Chave Acesso']
    ws.append(cab)
    for row in ws.iter_cols(min_col=1, max_col=len(cab), min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    for n in dados:
        ws.append([
            n['id'], n['numero'], n['serie'], n['nome_fornecedor'],
            n['cnpj_emitente'], n['valor_total'], n['status'],
            (n['data_emissao'] or '')[:10], n['chave_acesso'],
        ])
    ws2 = wb.create_sheet('Itens')
    ws2.append(['Nota ID', 'Código', 'Descrição', 'NCM', 'CFOP', 'Quantidade', 'Valor Unitário', 'Valor Total'])
    for row in ws2.iter_cols(min_col=1, max_col=8, min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
    for nota in notas:
        r = await db.execute(select(ItemNota).where(ItemNota.nota_id == nota.id))
        itens = r.scalars().all()
        for i in itens:
            ws2.append([nota.id, i.codigo, i.descricao, i.ncm, i.cfop, i.quantidade, i.valor_unitario, i.valor_total])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=notas_fiscais.xlsx'},
    )


@router.post('/{nota_id}/gerar-xml')
async def gerar_xml_nota(nota_id: int, db: AsyncSession = Depends(get_db)):
    from app.services.gerador_xml_service import gerar_xml
    try:
        caminho = await gerar_xml(nota_id, db)
        await db.commit()
        return FileResponse(caminho, media_type='application/xml', filename=os.path.basename(caminho))
    except ValueError as e:
        return {'erro': str(e)}


@router.get('/{nota_id}')
async def obter_nota(nota_id: int, db: AsyncSession = Depends(get_db)):
    svc = NotaService(db)
    nota = await svc.obter(nota_id)
    if not nota:
        return {'erro': 'Nota nao encontrada'}
    r = await db.execute(select(ItemNota).where(ItemNota.nota_id == nota_id))
    itens = r.scalars().all()
    return {
        'nota': NotaFiscalResponse.model_validate(nota),
        'itens': [ItemNotaResponse.model_validate(i) for i in itens],
    }

@router.post('', status_code=201)
async def criar_nota(dados: NotaFiscalCreate, db: AsyncSession = Depends(get_db)):
    svc = NotaService(db)
    nota = await svc.criar(dados)
    return NotaFiscalResponse.model_validate(nota)

@router.put('/{nota_id}')
async def atualizar_nota(nota_id: int, dados: NotaFiscalUpdate, db: AsyncSession = Depends(get_db)):
    svc = NotaService(db)
    nota = await svc.atualizar(nota_id, dados)
    if not nota:
        return {'erro': 'Nota nao encontrada'}
    return NotaFiscalResponse.model_validate(nota)

@router.delete('/{nota_id}')
async def deletar_nota(nota_id: int, db: AsyncSession = Depends(get_db)):
    svc = NotaService(db)
    ok = await svc.deletar(nota_id)
    return {'ok': ok}

@router.get('/{nota_id}/historico')
async def historico_nota(nota_id: int, db: AsyncSession = Depends(get_db)):
    svc = NotaService(db)
    historico = await svc.obter_historico(nota_id)
    return [HistoricoResponse.model_validate(h) for h in historico]

@router.get('/{nota_id}/comparacao')
async def comparar_nota(nota_id: int, db: AsyncSession = Depends(get_db)):
    from app.services.comparador_service import ComparadorService
    svc = ComparadorService(db)
    return await svc.comparar(nota_id)
