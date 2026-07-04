from typing import Optional
import openpyxl
from app.extractors import normalizar_valor, normalizar_data, extrair_chave_do_texto

def processar_xlsx(caminho: str) -> dict:
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            ws = wb.worksheets[0]
    except Exception as e:
        return {'erro': f'Erro ao ler XLSX: {e}'}

    linhas = []
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() if v is not None else '' for v in row]
        linhas.append(vals)

    dados = {
        'chave_acesso': '',
        'numero': '',
        'serie': '',
        'cnpj_emitente': '',
        'cpf_emitente': '',
        'nome_fornecedor': '',
        'data_emissao': '',
        'valor_total': 0.0,
        'tipo_documento': 'NF-e',
        'origem': 'xlsx',
        'itens': [],
    }

    dentro_tabela = False
    cabecalhos = []

    for i, linha in enumerate(linhas):
        texto_linha = ' '.join(linha)

        if not dados['cnpj_emitente']:
            for val in linha:
                m = __import__('re').search(r'\b(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})\b', val)
                if m:
                    dados['cnpj_emitente'] = m.group(1)
                    break

        if not dados['numero']:
            for val in linha:
                m = __import__('re').search(r'N[Ff][ea]?\s*[.:]?\s*(\d+)', val)
                if m:
                    dados['numero'] = m.group(1)
                    break

        if not dados['valor_total']:
            for val in linha:
                if 'TOTAL' in val.upper() or 'VALOR' in val.upper():
                    for j in range(len(linha)):
                        v = normalizar_valor(linha[j])
                        if v > 0:
                            dados['valor_total'] = v
                            break

        lc = [v.strip().upper() for v in linha if v.strip()]
        if any('COD' in c or 'PRODUTO' in c or 'DESCRI' in c for c in lc):
            dentro_tabela = True
            cabecalhos = [v.upper() for v in linha]
            continue

        if dentro_tabela:
            if any('TOTAL' in str(v).upper() for v in linha if v):
                dentro_tabela = False
                continue

            vals_nao_vazios = [v for v in linha if v.strip()]
            if len(vals_nao_vazios) >= 2:
                dados['itens'].append({
                    'codigo': linha[0] if len(linha) > 0 else '',
                    'descricao': linha[1] if len(linha) > 1 else '',
                    'ncm': '',
                    'cfop': '',
                    'quantidade': normalizar_valor(linha[3] if len(linha) > 3 and 'QUANT' in cabecalhos[3] else ''),
                    'valor_unitario': normalizar_valor(linha[4] if len(linha) > 4 else ''),
                    'valor_total': normalizar_valor(linha[5] if len(linha) > 5 else ''),
                })

    dados['qtd_itens'] = len(dados['itens'])
    return dados

def extrair_texto_xlsx(caminho: str) -> dict:
    try:
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active or wb.worksheets[0]
        linhas_texto = []
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() if v is not None else '' for v in row]
            linhas_texto.append(' | '.join(vals))
        return {'texto': '\n'.join(linhas_texto)}
    except Exception as e:
        return {'erro': str(e)}
